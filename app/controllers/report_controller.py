import calendar
import csv
import io
from datetime import date, timedelta

from flask import make_response, render_template, request, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.currency import CURRENCIES, format_currency
from app.models.transaction import TransactionModel


class ReportController:
    def _convert_chart_data(self, amounts_npr):
        currency = session.get("currency", "NPR")
        rate = CURRENCIES.get(currency, CURRENCIES["NPR"])["rate"]
        return [float(a or 0) * rate for a in amounts_npr]

    @staticmethod
    def _json_floats(values):
        return [float(v or 0) for v in values]

    def _resolve_date_range(self):
        range_type = request.args.get("range", "monthly")
        today = date.today()

        if range_type == "weekly":
            start = today - timedelta(days=6)
            end = today
            label = f"{start.isoformat()} – {end.isoformat()}"
        elif range_type == "yearly":
            start = date(today.year, 1, 1)
            end = today
            label = str(today.year)
        elif range_type == "custom":
            start_str = request.args.get("start_date", "")
            end_str = request.args.get("end_date", "")
            try:
                start = date.fromisoformat(start_str)
                end = date.fromisoformat(end_str)
                if start > end:
                    start, end = end, start
            except ValueError:
                start = date(today.year, today.month, 1)
                end = today
                range_type = "monthly"
            label = f"{start.isoformat()} – {end.isoformat()}"
        else:
            range_type = "monthly"
            month = request.args.get("month") or today.strftime("%Y-%m")
            year, mon = map(int, month.split("-"))
            start = date(year, mon, 1)
            last_day = calendar.monthrange(year, mon)[1]
            end = date(year, mon, last_day)
            if end > today:
                end = today
            label = month

        return range_type, start, end, label

    def index(self):
        user_id = session["user_id"]
        range_type, start, end, period_label = self._resolve_date_range()

        expense_by_category = TransactionModel.get_expense_by_category_range(
            user_id, start, end
        )
        summary = TransactionModel.get_summary_for_range(user_id, start, end)

        pie_labels = [r["category_name"] for r in expense_by_category]
        pie_data = self._json_floats(
            self._convert_chart_data([r["total"] for r in expense_by_category])
        )

        month = request.args.get("month") or date.today().strftime("%Y-%m")
        daily_rows = TransactionModel.get_daily_expenses(user_id, month)
        daily_labels = [str(r["day"])[-5:] for r in daily_rows]
        daily_data = self._json_floats(
            self._convert_chart_data([r["total"] for r in daily_rows])
        )

        six_month = TransactionModel.get_six_month_comparison(user_id)
        bar_months = [str(r["month"]) for r in six_month]
        bar_income = self._json_floats(
            self._convert_chart_data([r["income"] for r in six_month])
        )
        bar_expense = self._json_floats(
            self._convert_chart_data([r["expense"] for r in six_month])
        )

        if not bar_months:
            today = date.today()
            bar_months = []
            bar_income = []
            bar_expense = []
            for i in range(5, -1, -1):
                m = today.month - i
                y = today.year
                while m <= 0:
                    m += 12
                    y -= 1
                bar_months.append(f"{y:04d}-{m:02d}")
                bar_income.append(0)
                bar_expense.append(0)

        return render_template(
            "reports/index.html",
            range_type=range_type,
            period_label=period_label,
            month=month,
            start_date=start.isoformat(),
            end_date=end.isoformat(),
            summary=summary,
            pie_labels=pie_labels,
            pie_data=pie_data,
            daily_labels=daily_labels,
            daily_data=daily_data,
            bar_months=bar_months,
            bar_income=bar_income,
            bar_expense=bar_expense,
        )

    def export_csv(self):
        user_id = session["user_id"]
        transactions = TransactionModel.get_all_for_export(user_id)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            ["Date", "Type", "Category", "Amount (NPR)", "Description", "Notes"]
        )
        for t in transactions:
            writer.writerow(
                [
                    str(t["date"]),
                    t["type"].capitalize(),
                    t.get("category_name") or "",
                    float(t["amount"]),
                    t.get("description") or "",
                    t.get("notes") or "",
                ]
            )

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = (
            "attachment; filename=expensex_report.csv"
        )
        return response

    def export_pdf(self):
        user_id = session["user_id"]
        transactions = TransactionModel.get_all_for_export(user_id)
        currency = session.get("currency", "NPR")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("ExpenseX — Transaction Report", styles["Title"]))
        elements.append(Spacer(1, 12))

        table_data = [
            ["Date", "Type", "Category", "Amount", "Description", "Notes"],
        ]
        for t in transactions:
            amount_str = format_currency(t["amount"], currency)
            table_data.append(
                [
                    str(t["date"]),
                    t["type"].capitalize(),
                    t.get("category_name") or "—",
                    amount_str,
                    t.get("description") or "—",
                    (t.get("notes") or "—")[:80],
                ]
            )

        if len(table_data) == 1:
            table_data.append(["—", "—", "—", "—", "No transactions", "—"])

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D1117")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#4F7EFF")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#131922")),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#E8EEF8")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1E2B3F")),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = (
            "attachment; filename=expensex_report.pdf"
        )
        return response

    def export_excel(self):
        user_id = session["user_id"]
        transactions = TransactionModel.get_all_for_export(user_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ["Date", "Type", "Category", "Amount (NPR)", "Description", "Notes"]
        header_fill = PatternFill(
            start_color="0D1117", end_color="0D1117", fill_type="solid"
        )
        header_font = Font(bold=True, color="4F7EFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, t in enumerate(transactions, 2):
            ws.cell(row=row_idx, column=1, value=str(t["date"]))
            ws.cell(row=row_idx, column=2, value=t["type"].capitalize())
            ws.cell(row=row_idx, column=3, value=t.get("category_name") or "—")
            ws.cell(row=row_idx, column=4, value=float(t["amount"]))
            ws.cell(row=row_idx, column=5, value=t.get("description") or "—")
            ws.cell(row=row_idx, column=6, value=t.get("notes") or "—")

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = make_response(buffer.read())
        response.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = (
            "attachment; filename=expensex_report.xlsx"
        )
        return response


#this code is done and its working perfectly 
import io
import math

from flask import flash, make_response, redirect, render_template, request, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.auth import validate_csrf
from app.database import log_activity
from app.models.admin import AdminModel
from app.models.user import UserModel


class AdminController:
    def _paginate(self, total, page, per_page):
        total_pages = max(1, math.ceil(total / per_page))
        page = max(1, min(page, total_pages))
        return {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "has_prev": page > 1,
            "has_next": page < total_pages,
        }

    def dashboard(self):
        stats = AdminModel.get_platform_stats()
        activity = AdminModel.get_activity_feed(15)
        signups = AdminModel.get_signups_by_month(6)
        signup_labels = [r["month"] for r in signups]
        signup_data = [int(r["count"]) for r in signups]
        top_users = AdminModel.get_top_active_users(5)
        return render_template(
            "admin/dashboard.html",
            stats=stats,
            activity=activity,
            signup_labels=signup_labels,
            signup_data=signup_data,
            top_users=top_users,
        )

    def users(self):
        page = request.args.get("page", 1, type=int)
        per_page = AdminModel.PER_PAGE
        total = AdminModel.count_users()
        users = AdminModel.get_all_users(page, per_page)
        pagination = self._paginate(total, page, per_page)
        return render_template(
            "admin/users.html", users=users, pagination=pagination
        )

    def user_detail(self, user_id):
        user = AdminModel.get_user_by_id(user_id)
        if not user:
            flash("User not found.", "danger")
            return redirect(url_for("admin.users"))

        totals = UserModel.get_totals(user_id)
        transactions = AdminModel.get_user_transactions(user_id)
        txn_count = len(transactions)
        balance = totals["income"] - totals["expense"]

        return render_template(
            "admin/user_detail.html",
            user=user,
            transactions=transactions,
            totals=totals,
            balance=balance,
            txn_count=txn_count,
        )

    def toggle_status(self, user_id):
        if not validate_csrf():
            flash("Invalid security token.", "danger")
            return redirect(url_for("admin.users"))

        if user_id == session.get("user_id"):
            flash("You cannot deactivate your own account.", "danger")
            return redirect(url_for("admin.users"))

        result = AdminModel.toggle_user_status(user_id)
        if result[1]:
            log_activity(
                session["user_id"],
                session.get("user_name", ""),
                "User status changed",
                result[1],
            )
        flash("User status updated.", "success")
        return redirect(url_for("admin.users"))

    def change_role(self, user_id):
        if not validate_csrf():
            flash("Invalid security token.", "danger")
            return redirect(url_for("admin.users"))

        role = request.form.get("role", "user")
        detail = AdminModel.change_user_role(user_id, role)
        if detail:
            log_activity(
                session["user_id"],
                session.get("user_name", ""),
                "Role changed",
                detail,
            )
        flash("User role updated.", "success")
        return redirect(url_for("admin.users"))

    def delete_user(self, user_id):
        if not validate_csrf():
            flash("Invalid security token.", "danger")
            return redirect(url_for("admin.users"))

        if user_id == session.get("user_id"):
            flash("You cannot delete your own account.", "danger")
            return redirect(url_for("admin.users"))

        ok, detail = AdminModel.delete_user(user_id)
        if not ok:
            if detail == "last_admin":
                flash("Cannot delete the last admin account.", "danger")
            elif detail == "not_found":
                flash("User not found.", "danger")
            else:
                flash("Could not delete user.", "danger")
            return redirect(url_for("admin.users"))

        log_activity(
            session["user_id"],
            session.get("user_name", ""),
            "User deleted",
            detail,
        )
        flash(f"User “{detail}” deleted.", "success")
        return redirect(url_for("admin.users"))

    def audit_log(self):
        month = request.args.get("month")
        user_id = request.args.get("user_id", type=int)
        txn_type = request.args.get("type", "")
        page = request.args.get("page", 1, type=int)
        per_page = 30
        total = AdminModel.count_all_transactions(month, user_id, txn_type or None)
        transactions = AdminModel.get_all_transactions(
            month, user_id, txn_type or None, page, per_page
        )
        users = AdminModel.get_all_users(page=1, per_page=500)
        pagination = self._paginate(total, page, per_page)
        return render_template(
            "admin/audit.html",
            transactions=transactions,
            pagination=pagination,
            month=month or "",
            user_id=user_id,
            txn_type=txn_type,
            users=users,
        )

    def audit_export_excel(self):
        month = request.args.get("month")
        user_id = request.args.get("user_id", type=int)
        txn_type = request.args.get("type", "")
        transactions = AdminModel.get_all_transactions_export(
            month, user_id, txn_type or None
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Log"
        headers = [
            "Date",
            "User",
            "Type",
            "Amount (NPR)",
            "Category",
            "Description",
        ]
        header_fill = PatternFill(
            start_color="0D1117", end_color="0D1117", fill_type="solid"
        )
        header_font = Font(bold=True, color="4F7EFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, t in enumerate(transactions, 2):
            ws.cell(row=row_idx, column=1, value=str(t["date"]))
            ws.cell(row=row_idx, column=2, value=t.get("user_name") or "")
            ws.cell(row=row_idx, column=3, value=t["type"].capitalize())
            ws.cell(row=row_idx, column=4, value=float(t["amount"]))
            ws.cell(row=row_idx, column=5, value=t.get("category_name") or "")
            ws.cell(row=row_idx, column=6, value=t.get("description") or "")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = make_response(buffer.read())
        response.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = (
            "attachment; filename=expensex_audit.xlsx"
        )
        return response

    def global_categories(self):
        categories = AdminModel.get_global_categories()
        return render_template("admin/categories.html", categories=categories)

    def add_global_category(self):
        if not validate_csrf():
            flash("Invalid security token.", "danger")
            return redirect(url_for("admin.global_categories"))

        name = request.form.get("name", "").strip()
        cat_type = request.form.get("type", "expense")
        if cat_type not in ("income", "expense"):
            cat_type = "expense"

        if not name:
            flash("Category name is required.", "danger")
            return redirect(url_for("admin.global_categories"))

        if AdminModel.global_category_exists(name, cat_type):
            flash("Global category already exists.", "danger")
            return redirect(url_for("admin.global_categories"))

        AdminModel.create_global_category(name, cat_type)
        flash("Global category created.", "success")
        return redirect(url_for("admin.global_categories"))

    def delete_global_category(self, cat_id):
        if not validate_csrf():
            flash("Invalid security token.", "danger")
            return redirect(url_for("admin.global_categories"))

        AdminModel.delete_global_category(cat_id)
        flash("Global category deleted.", "success")
        return redirect(url_for("admin.global_categories"))
